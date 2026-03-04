#!/usr/bin/env python3
import argparse
import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

CONCEPT_HEADER = [
    'conceptId',
    'name',
    'description',
    'conceptTypeName',
    'rootConceptId',
    'rootConceptName',
    'partOfConceptId',
    'partOfName',
    'partOrder',
    'referenceToConceptId',
    'referenceToName',
]

CONCEPT_TYPE_HEADER = ['name', 'description', 'partOfName', 'partOrder', 'referenceToName']


def normalize_text(value: object) -> str:
    if value is None:
        return ''
    text = str(value)
    text = text.replace('\xa0', ' ')
    text = re.sub(r'\s+', ' ', text)
    return text.strip()


def normalize_key(value: str) -> str:
    return normalize_text(value).lower()


def split_list_cell(value: object) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []

    items = [normalize_text(item) for item in text.replace('\n', ',').split(',')]
    return [item for item in items if item]


@dataclass
class ConceptRow:
    name: str
    description: str
    concept_type_name: str
    root_concept_name: str = ''
    part_of_name: str = ''
    part_order: str = ''
    reference_to_name: str = ''

    def to_csv_row(self) -> list[str]:
        return [
            '',
            self.name,
            self.description,
            self.concept_type_name,
            '',
            self.root_concept_name,
            '',
            self.part_of_name,
            self.part_order,
            '',
            self.reference_to_name,
        ]


def write_csv(path: Path, header: list[str], rows: Iterable[list[str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open('w', newline='', encoding='utf-8') as handle:
        writer = csv.writer(handle)
        writer.writerow(header)
        for row in rows:
            writer.writerow(row)


def extract_capability_rows(workbook_path: Path) -> list[ConceptRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook['Capability Map']

    rows: list[ConceptRow] = []
    level_stack: dict[int, str] = {}
    root_by_name: dict[str, str] = {}

    blank_streak = 0
    for row_index, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), start=3):
        _tier_raw, level_raw, capability_name, capability_definition, *_ = list(row) + [None] * 8
        name = normalize_text(capability_name)
        definition = normalize_text(capability_definition)

        if not name and not normalize_text(level_raw):
            blank_streak += 1
            if blank_streak >= 25:
                break
            continue

        blank_streak = 0

        if not name:
            continue

        if not isinstance(level_raw, int):
            try:
                level = int(str(level_raw).strip())
            except Exception:
                continue
        else:
            level = level_raw

        level_stack[level] = name
        for existing_level in list(level_stack.keys()):
            if existing_level > level:
                del level_stack[existing_level]

        if level == 1:
            root_name = name
            root_by_name[name] = name
            rows.append(
                ConceptRow(
                    name=name,
                    description=definition,
                    concept_type_name='Capability',
                )
            )
            continue

        parent_name = level_stack.get(level - 1, '')
        root_name = root_by_name.get(level_stack.get(1, ''), level_stack.get(1, ''))
        if not parent_name:
            continue

        rows.append(
            ConceptRow(
                name=name,
                description=definition,
                concept_type_name='Capability',
                root_concept_name=root_name,
                part_of_name=parent_name,
                part_order='',
            )
        )

    workbook.close()
    return rows


def extract_value_stream_rows(workbook_path: Path) -> list[ConceptRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    inventory = workbook['Value Stream Inventory']

    rows: list[ConceptRow] = []
    value_streams: list[tuple[str, str]] = []

    blank_streak = 0
    for row in inventory.iter_rows(min_row=3, values_only=True):
        stream_name = normalize_text(row[0] if len(row) > 0 else None)
        stream_description = normalize_text(row[1] if len(row) > 1 else None)
        if not stream_name and not stream_description:
            blank_streak += 1
            if blank_streak >= 25:
                break
            continue

        blank_streak = 0
        if not stream_name:
            continue
        value_streams.append((stream_name, stream_description))

    for stream_name, stream_description in value_streams:
        rows.append(
            ConceptRow(
                name=stream_name,
                description=stream_description,
                concept_type_name='Value Stream',
            )
        )

        if stream_name not in workbook.sheetnames:
            continue

        sheet = workbook[stream_name]
        stage_order = 1
        stage_blank_streak = 0
        for row in sheet.iter_rows(min_row=3, values_only=True):
            stage_name = normalize_text(row[1] if len(row) > 1 else None)
            stage_description = normalize_text(row[2] if len(row) > 2 else None)

            if not stage_name:
                stage_blank_streak += 1
                if stage_blank_streak >= 20:
                    break
                continue

            stage_blank_streak = 0
            if stage_name.lower().startswith('return to the value stream inventory'):
                continue

            rows.append(
                ConceptRow(
                    name=stage_name,
                    description=stage_description,
                    concept_type_name='Value Stream Stage',
                    root_concept_name=stream_name,
                    part_of_name=stream_name,
                    part_order=str(stage_order),
                )
            )
            stage_order += 1

    workbook.close()
    return rows


def extract_information_rows(workbook_path: Path) -> list[ConceptRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook['Information Map']

    rows: list[ConceptRow] = []
    extracted_rows: list[tuple[str, str, str, list[str], list[str], list[str]]] = []
    information_names: set[str] = set()

    blank_streak = 0
    for row in sheet.iter_rows(min_row=3, values_only=True):
        concept_name = normalize_text(row[0] if len(row) > 0 else None)
        concept_category = normalize_text(row[1] if len(row) > 1 else None)
        concept_definition = normalize_text(row[2] if len(row) > 2 else None)
        concept_types = split_list_cell(row[3] if len(row) > 3 else None)
        related_concepts = split_list_cell(row[4] if len(row) > 4 else None)
        concept_states = split_list_cell(row[5] if len(row) > 5 else None)

        if not concept_name:
            blank_streak += 1
            if blank_streak >= 25:
                break
            continue

        blank_streak = 0

        if ':' in concept_name and not concept_category and not concept_definition:
            continue

        extracted_rows.append(
            (
                concept_name,
                concept_category,
                concept_definition,
                concept_types,
                related_concepts,
                concept_states,
            )
        )
        information_names.add(concept_name)

    information_name_keys = {normalize_key(name) for name in information_names}

    for concept_name, concept_category, concept_definition, concept_types, related_concepts, concept_states in extracted_rows:
        category_prefix = f'Category: {concept_category}. ' if concept_category else ''
        rows.append(
            ConceptRow(
                name=concept_name,
                description=f'{category_prefix}{concept_definition}'.strip(),
                concept_type_name='Information Concept',
            )
        )

        seen_types: set[str] = set()
        for index, type_name in enumerate(concept_types, start=1):
            type_key = normalize_key(type_name)
            if not type_key or type_key in seen_types:
                continue
            seen_types.add(type_key)
            rows.append(
                ConceptRow(
                    name=type_name,
                    description=f'Type for {concept_name}',
                    concept_type_name='Information Concept Type',
                    root_concept_name=concept_name,
                    part_of_name=concept_name,
                    part_order=str(index),
                )
            )

        seen_states: set[str] = set()
        for index, state_name in enumerate(concept_states, start=1):
            state_key = normalize_key(state_name)
            if not state_key or state_key in seen_states:
                continue
            seen_states.add(state_key)
            rows.append(
                ConceptRow(
                    name=state_name,
                    description=f'State for {concept_name}',
                    concept_type_name='Information Concept State',
                    root_concept_name=concept_name,
                    part_of_name=concept_name,
                    part_order=str(index),
                )
            )

        seen_relations: set[str] = set()
        relation_order = 1
        for related_name in related_concepts:
            related_key = normalize_key(related_name)
            if not related_key or related_key in seen_relations:
                continue
            if related_key not in information_name_keys:
                continue
            seen_relations.add(related_key)

            link_name = f'{concept_name} -> {related_name}'
            rows.append(
                ConceptRow(
                    name=link_name,
                    description=f'Related information concept link from {concept_name} to {related_name}',
                    concept_type_name='Information Concept Link',
                    root_concept_name=concept_name,
                    part_of_name=concept_name,
                    part_order=str(relation_order),
                    reference_to_name=related_name,
                )
            )
            relation_order += 1

    workbook.close()
    return rows


def extract_stakeholder_rows(workbook_path: Path) -> list[ConceptRow]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook['Stakeholder Map']

    grouped: dict[str, dict[str, object]] = {}

    blank_streak = 0
    for row in sheet.iter_rows(min_row=3, values_only=True):
        stakeholder_type = normalize_text(row[0] if len(row) > 0 else None)
        stakeholder_category = normalize_text(row[1] if len(row) > 1 else None)
        stakeholder_name = normalize_text(row[2] if len(row) > 2 else None)
        stakeholder_description = normalize_text(row[3] if len(row) > 3 else None)

        if not stakeholder_name:
            blank_streak += 1
            if blank_streak >= 25:
                break
            continue

        blank_streak = 0

        key = normalize_key(stakeholder_name)
        if key not in grouped:
            grouped[key] = {
                'name': stakeholder_name,
                'description': stakeholder_description,
                'types': set(),
                'categories': set(),
            }

        entry = grouped[key]
        if stakeholder_type:
            entry['types'].add(stakeholder_type)
        if stakeholder_category:
            entry['categories'].add(stakeholder_category)
        if not entry['description'] and stakeholder_description:
            entry['description'] = stakeholder_description

    rows: list[ConceptRow] = []
    for _, entry in sorted(grouped.items(), key=lambda item: str(item[1]['name']).lower()):
        type_text = ', '.join(sorted(entry['types']))
        category_text = ', '.join(sorted(entry['categories']))
        base_description = normalize_text(entry['description'])
        prefixes = []
        if category_text:
            prefixes.append(f'Category: {category_text}')
        if type_text:
            prefixes.append(f'Type: {type_text}')
        prefix_text = '. '.join(prefixes)
        final_description = base_description
        if prefix_text:
            final_description = f'{prefix_text}. {base_description}'.strip()

        rows.append(
            ConceptRow(
                name=str(entry['name']),
                description=final_description,
                concept_type_name='Stakeholder',
            )
        )

    workbook.close()
    return rows


def build_concept_types_rows() -> list[list[str]]:
    return [
        ['Capability', 'Government business capability (supports decomposition)', 'Capability', '1', ''],
        ['Value Stream', 'End-to-end value delivery perspective', '', '', ''],
        ['Value Stream Stage', 'Stage within a value stream', 'Value Stream', '1', ''],
        ['Information Concept', 'Business information object', '', '', ''],
        ['Information Concept Type', 'Classification of an information concept', 'Information Concept', '1', ''],
        ['Information Concept State', 'State/status of an information concept', 'Information Concept', '2', ''],
        ['Information Concept Link', 'Bridge concept for related information concepts', 'Information Concept', '3', 'Information Concept'],
        ['Stakeholder', 'Person or organization role in value delivery', '', '', ''],
    ]


def dedupe_concept_rows(rows: list[ConceptRow]) -> list[ConceptRow]:
    deduped: list[ConceptRow] = []
    seen: set[tuple[str, str, str, str, str]] = set()

    for row in rows:
        key = (
            normalize_key(row.concept_type_name),
            normalize_key(row.root_concept_name),
            normalize_key(row.name),
            normalize_key(row.part_of_name),
            normalize_key(row.reference_to_name),
        )
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)

    return deduped


def main() -> int:
    parser = argparse.ArgumentParser(description='Build reference model import package CSVs from a BAG workbook.')
    parser.add_argument('--source-xlsx', required=True, help='Path to source workbook (.xlsx).')
    parser.add_argument(
        '--out-dir',
        required=True,
        help='Output directory for package CSV files (for example reference-models/government/v5.0).',
    )
    args = parser.parse_args()

    source_path = Path(args.source_xlsx).resolve()
    out_dir = Path(args.out_dir).resolve()

    if not source_path.exists():
        raise FileNotFoundError(f'Source workbook not found: {source_path}')

    capability_rows = dedupe_concept_rows(extract_capability_rows(source_path))
    value_stream_rows = dedupe_concept_rows(extract_value_stream_rows(source_path))
    information_rows = dedupe_concept_rows(extract_information_rows(source_path))
    stakeholder_rows = dedupe_concept_rows(extract_stakeholder_rows(source_path))

    write_csv(out_dir / '01-concept-types.csv', CONCEPT_TYPE_HEADER, build_concept_types_rows())
    write_csv(out_dir / '02-concepts-capability.csv', CONCEPT_HEADER, [row.to_csv_row() for row in capability_rows])
    write_csv(out_dir / '03-concepts-value-stream.csv', CONCEPT_HEADER, [row.to_csv_row() for row in value_stream_rows])
    write_csv(out_dir / '04-concepts-information.csv', CONCEPT_HEADER, [row.to_csv_row() for row in information_rows])
    write_csv(out_dir / '05-concepts-stakeholder.csv', CONCEPT_HEADER, [row.to_csv_row() for row in stakeholder_rows])

    print('Reference model package files generated:')
    print(f'- {out_dir / "01-concept-types.csv"}')
    print(f'- {out_dir / "02-concepts-capability.csv"} ({len(capability_rows)} rows)')
    print(f'- {out_dir / "03-concepts-value-stream.csv"} ({len(value_stream_rows)} rows)')
    print(f'- {out_dir / "04-concepts-information.csv"} ({len(information_rows)} rows)')
    print(f'- {out_dir / "05-concepts-stakeholder.csv"} ({len(stakeholder_rows)} rows)')

    return 0


if __name__ == '__main__':
    raise SystemExit(main())
